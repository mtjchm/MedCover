#!/usr/bin/env python3
"""
MedCover – Google Sheets Event & User Extraction Script (v2)
============================================================
Reads a downloaded .xlsx export of the "Dozory" Google Sheet and produces
a JSON file suitable for pasting into the MedCover import page.

This v2 script extracts both events (from the "Dozory" sheet) and users
(from "Lidi" cross-referenced with "Dozory") in a single pass.

Usage
-----
    python scripts/import_events.py [--input <file.xlsx>] [--output <file.json>]

Defaults:
    --input   Dozory 2026.xlsx  (relative to CWD)
    --output  import_events.json (relative to CWD)

Requirements
------------
    pip install openpyxl

Column mapping (GS "Dozory" sheet, row 2 = headers, data from row 3):
    A (0)  Akce                  → name
    B (1)  Datum                 → date  (datetime.datetime)
    C (2)  Místo konání          → location
    D (3)  Vozidlo/stan          → description snippet
    E (4)  Osvěta pp / dozor     → description snippet
    F (5)  kontakt pořadatel     → contact_person
    G (6)  Začátek               → start_time  (datetime.time or None)
    H (7)  Konec                 → end_time    (datetime.time or None)
    I (8)  Hlídky zdravotníků    → (ignored)
    J (9)  Počet ošetření        → (ignored)
    K (10) Placený               → paid  (bool)
    L (11) Doba trvání           → (ignored – computed by DB)
    M (12) Zodpovědný zdravotník → responsible_person (converted to Firstname Lastname)
    N+ (13+) Přihlášení         → signups list (converted names) + description snippet

Column mapping (GS "Lidi" sheet, data from row 2):
    A (0)  Jméno      → gs_name (Lastname Firstname format)
    B (1)  Telefon    → phone
    C (2)  E-mail     → email
    D (3)  Zdravotník → is_zdravotnik (bool)

Output JSON schema (v2):
    {
        "version": 2,
        "users": [
            {
                "gs_name":       str,   # original "Lastname Firstname" from GS
                "name":          str,   # converted "Firstname Lastname"
                "email":         str or null,
                "phone":         str or null,
                "is_zdravotnik": bool
            }
        ],
        "events": [
            {
                "name":               str,
                "date":               "YYYY-MM-DD",
                "start_time":         "HH:MM" or null,
                "end_time":           "HH:MM" or null,
                "location":           str or null,
                "paid":               bool,
                "responsible_person": str or null,  # "Firstname Lastname"
                "contact_person":     str or null,
                "description":        str,
                "time_missing":       bool,
                "signups":            list[str]      # "Firstname Lastname" of col N+ people
            }
        ]
    }

Notes
-----
- Past events (before today) are filtered out.
- Duplicate event names get the date appended: "Vítání občánků 22.1."
- Names are kept as-is from GS ("Surname Firstname"), matching MedCover's convention.
- Users are extracted from all names in Dozory (cols M+N+) cross-referenced with Lidi.
- People not found in Lidi will have null email and phone — email must be filled manually.
- Spot count: ≤3 people → standard 3-spot pattern; >3 people → dynamic count.
- The script is idempotent: running it multiple times produces the same output.
  Duplicate detection is handled by the web app at import time.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any


def _fmt_time(t: Any) -> str | None:
    """Return 'HH:MM' string from a datetime.time object, or None."""
    if t is None:
        return None
    try:
        return t.strftime("%H:%M")
    except AttributeError:
        return None


def _fmt_date(dt: Any) -> str | None:
    """Return 'YYYY-MM-DD' string from a datetime.datetime object, or None."""
    if not isinstance(dt, datetime):
        return None
    return dt.strftime("%Y-%m-%d")


def _reverse_name(gs_name: str) -> str:
    """Convert 'Lastname Firstname [Middle]' (GS format) to 'Firstname [Middle] Lastname'.

    Examples:
        "Balhar Lumír"        → "Lumír Balhar"
        "Svobodová K. Zuzana" → "K. Zuzana Svobodová"
    """
    parts = gs_name.strip().split(None, 1)
    if len(parts) == 2:
        return f"{parts[1]} {parts[0]}"
    return gs_name.strip()


def _is_valid_name(name: str) -> bool:
    """Return True if the string looks like a real person's name (contains at least one letter)."""
    return bool(name) and any(c.isalpha() for c in name)


def _build_description(
    vehicle: str | None,
    event_type: str | None,
    contact: str | None,
    signups: list[str],
    time_missing: bool,
) -> str:
    """Assemble the description field from GS columns that have no dedicated field."""
    parts: list[str] = []
    if event_type:
        parts.append(f"Typ: {event_type.strip()}")
    if vehicle:
        parts.append(f"Vozidlo/stan: {vehicle.strip()}")
    if contact:
        parts.append(f"Kontakt pořadatel: {contact.strip()}")
    if signups:
        names = ", ".join(n.strip() for n in signups if n and str(n).strip())
        if names:
            parts.append(f"Přihlášení (import z GS): {names}")
    if time_missing:
        parts.append(
            "UPOZORNĚNÍ: Čas akce nebyl v importních datech uveden. "
            "Akce byla vytvořena bez pozic. Doplňte čas a přidejte pozice ručně."
        )
    return " | ".join(parts)


def extract(wb: Any, cutoff: date | None = None) -> list[dict[str, Any]]:
    """Read the Dozory sheet and return a list of event dicts.

    Args:
        wb:     Opened openpyxl workbook.
        cutoff: Only include events with date >= cutoff.  Defaults to today.

    Returns:
        List of event dicts in the v2 interchange JSON schema.
    """
    if cutoff is None:
        cutoff = date.today()

    ws = wb["Dozory"]

    # Count occurrences of each name among future events so we know
    # which names need the date suffix to stay unique.
    name_counts: dict[str, int] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[0]
        dt = row[1]
        if not name or not isinstance(dt, datetime):
            continue
        if dt.date() < cutoff:
            continue
        name_counts[str(name)] = name_counts.get(str(name), 0) + 1

    events: list[dict[str, Any]] = []
    seen_names: dict[str, int] = {}  # track how many times we've used a name+suffix

    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[0]
        dt = row[1]

        # Skip rows without a name or a valid date
        if not name or not isinstance(dt, datetime):
            continue

        # Filter past events
        if dt.date() < cutoff:
            continue

        name = str(name).strip()
        raw_date = dt.date()

        # --- Name uniqueness ---
        if name_counts.get(name, 1) > 1:
            suffix = raw_date.strftime("%-d.%-m.")
            unique_name = f"{name} {suffix}"
        else:
            unique_name = name

        # If even after appending date there's a clash, add a counter.
        if unique_name in seen_names:
            seen_names[unique_name] += 1
            unique_name = f"{unique_name} ({seen_names[unique_name]})"
        else:
            seen_names[unique_name] = 1

        # --- Times ---
        start_time = row[6]
        end_time = row[7]
        time_missing = start_time is None

        # Treat midnight end as "end not specified"
        if end_time is not None:
            from datetime import time as _time  # noqa: PLC0415
            if end_time == _time(0, 0):
                end_time = None

        # --- Responsible person (GS "Surname Firstname" = MedCover format, no conversion) ---
        responsible_person_gs = row[12]
        if responsible_person_gs is not None:
            rp_str = str(responsible_person_gs).strip()
            responsible_person: str | None = rp_str if _is_valid_name(rp_str) else None
        else:
            responsible_person = None

        # --- Signups (col N onwards, already "Surname Firstname") ---
        signup_gs_names = [
            str(v).strip()
            for v in row[13:]
            if v is not None and _is_valid_name(str(v).strip())
        ]
        signup_converted = list(signup_gs_names)

        description = _build_description(
            vehicle=row[3],
            event_type=row[4],
            contact=row[5],
            signups=signup_gs_names,
            time_missing=time_missing,
        )

        paid_raw = row[10]
        # GS stores as bool True/False; some rows may have None or 0/1
        if isinstance(paid_raw, bool):
            paid = paid_raw
        elif isinstance(paid_raw, (int, float)):
            paid = bool(paid_raw)
        else:
            paid = False

        location = row[2]
        if location is not None:
            location = str(location).strip() or None

        contact_person = row[5]
        if contact_person is not None:
            contact_person = str(contact_person).strip() or None

        events.append(
            {
                "name": unique_name,
                "date": raw_date.strftime("%Y-%m-%d"),
                "start_time": _fmt_time(start_time),
                "end_time": _fmt_time(end_time),
                "location": location,
                "paid": paid,
                "responsible_person": responsible_person,
                "contact_person": contact_person,
                "description": description,
                "time_missing": time_missing,
                "signups": signup_converted,
            }
        )

    return events


def _load_lidi_lookup(wb: Any) -> dict[str, dict[str, Any]]:
    """Load the 'Lidi' sheet into a dict keyed by original GS name (Lastname Firstname)."""
    ws = wb["Lidi"]
    lookup: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        gs_name = str(row[0]).strip()
        if not _is_valid_name(gs_name):
            continue
        phone_raw = row[1]
        email_raw = row[2]
        zdravotnik_raw = row[3]
        phone = str(phone_raw).strip() if phone_raw is not None else ""
        email = str(email_raw).strip() if email_raw is not None else ""

        if isinstance(zdravotnik_raw, bool):
            is_zdravotnik = zdravotnik_raw
        elif isinstance(zdravotnik_raw, str):
            is_zdravotnik = zdravotnik_raw.strip().lower() in ("true", "ano", "1", "yes")
        else:
            is_zdravotnik = bool(zdravotnik_raw) if zdravotnik_raw is not None else False

        lookup[gs_name] = {
            "phone": phone or None,
            "email": email or None,
            "is_zdravotnik": is_zdravotnik,
        }
    return lookup


def extract_users(wb: Any, cutoff: date | None = None) -> list[dict[str, Any]]:
    """Extract unique users from Dozory who appear as RP (col M) or signups (cols N+).

    Cross-references with the Lidi sheet to get email, phone, and zdravotník flag.
    Names are kept as-is from GS ("Surname Firstname"), matching MedCover convention.

    Args:
        wb:     Opened openpyxl workbook.
        cutoff: Only consider events on or after this date.  Defaults to today.

    Returns:
        Sorted list of user dicts, one per unique person found in Dozory.
    """
    if cutoff is None:
        cutoff = date.today()

    lidi = _load_lidi_lookup(wb)
    ws = wb["Dozory"]
    seen: set[str] = set()
    users: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        dt = row[1]
        if not isinstance(dt, datetime) or dt.date() < cutoff:
            continue

        # Collect all names: col M (RP) + cols N+ (signups)
        all_gs_names: list[str] = []
        if row[12]:
            n = str(row[12]).strip()
            if _is_valid_name(n):
                all_gs_names.append(n)
        for v in row[13:]:
            if v:
                n = str(v).strip()
                if _is_valid_name(n) and n not in all_gs_names:
                    all_gs_names.append(n)

        for gs_name in all_gs_names:
            if gs_name in seen:
                continue
            seen.add(gs_name)

            lidi_info = lidi.get(gs_name)
            users.append({
                "gs_name": gs_name,
                "name": gs_name,  # GS "Surname Firstname" = MedCover convention
                "email": lidi_info["email"] if lidi_info else None,
                "phone": lidi_info["phone"] if lidi_info else None,
                "is_zdravotnik": lidi_info["is_zdravotnik"] if lidi_info else False,
            })

    users.sort(key=lambda u: u["name"])
    return users


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Extract future events and users from a Google Sheets .xlsx export to JSON (v2)."
        )
    )
    parser.add_argument(
        "--input",
        default="Dozory 2026.xlsx",
        help="Path to the .xlsx file (default: 'Dozory 2026.xlsx')",
    )
    parser.add_argument(
        "--output",
        default="import_events.json",
        help="Output JSON file path (default: 'import_events.json'). Use '-' for stdout.",
    )
    parser.add_argument(
        "--cutoff",
        default=None,
        help="Only include events on or after this date (YYYY-MM-DD). Default: today.",
    )
    args = parser.parse_args()

    cutoff: date | None = None
    if args.cutoff:
        try:
            cutoff = date.fromisoformat(args.cutoff)
        except ValueError:
            print(f"ERROR: Invalid --cutoff date: {args.cutoff}", file=sys.stderr)
            sys.exit(1)

    xlsx_path = Path(args.input)
    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    try:
        import openpyxl  # noqa: PLC0415
    except ImportError:
        print(
            "ERROR: openpyxl is not installed.  Run:  pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    events = extract(wb, cutoff=cutoff)
    users = extract_users(wb, cutoff=cutoff)

    output = {
        "version": 2,
        "users": users,
        "events": events,
    }
    output_json = json.dumps(output, ensure_ascii=False, indent=2)

    if args.output == "-":
        print(output_json)
    else:
        out_path = Path(args.output)
        out_path.write_text(output_json, encoding="utf-8")
        print(
            f"✓ {len(users)} osob, {len(events)} akcí → {out_path}",
            file=sys.stderr,
        )


if __name__ == "__main__":
    main()
