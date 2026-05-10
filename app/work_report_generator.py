"""
Výkaz práce (employee work report) xlsx generator.

Generates a single-sheet openpyxl workbook that matches the layout of the
legacy Google-Sheets "Dozory YYYY.xlsx" monthly report used by Czech Red
Cross members to document worked hours for DPP payroll purposes.

Public entry point::

    path = generate_work_report(user, year, month)

The file is written to  instance/work_report/<user_id>/<year>-<MM>.xlsx
and is overwritten on each call.  Callers are responsible for serving
the file and scheduling cleanup (files older than 1 day should be removed).
"""
from __future__ import annotations

import calendar
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING

import holidays
import sqlalchemy as sa
from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:
    from app.models import UserAccount

# ── Czech locale constants ────────────────────────────────────────────────────

CZ_MONTH_NAMES = [
    "", "Leden", "Únor", "Březen", "Duben", "Květen", "Červen",
    "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec",
]

CZ_WEEKDAY_ABBR = ["PO", "ÚT", "ST", "ČT", "PÁ", "SO", "NE"]  # Mon=0 … Sun=6

# ── Colours ──────────────────────────────────────────────────────────────────

_BLUE_FILL = PatternFill("solid", fgColor="FF99CCFF")   # column header row
_CYAN_FILL = PatternFill("solid", fgColor="FFCCFFFF")   # info block labels
_YELLOW_FILL = PatternFill("solid", fgColor="FFFFFF00")  # public holiday
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFFFF")  # normal day
_RED_FONT = Font(name="Calibri", size=10, color="FFFF0000")   # weekend day name
_STD_FONT = Font(name="Calibri", size=10)
_BOLD_FONT = Font(name="Calibri", size=10, bold=True)

# ── Borders ──────────────────────────────────────────────────────────────────


def _side(style: str) -> Side:
    return Side(border_style=style)


_HDR_BORDER_A = Border(
    left=_side("medium"), right=_side("thin"),
    top=_side("medium"), bottom=_side("medium"),
)
_HDR_BORDER_B = Border(
    left=_side("medium"), right=_side("thin"),
    top=_side("medium"),
)
_HDR_BORDER_C = Border(
    left=_side("medium"), right=_side("thin"),
    top=_side("medium"), bottom=_side("medium"),
)
_HDR_BORDER_D = Border(
    left=_side("thin"), right=_side("medium"),
    top=_side("medium"), bottom=_side("medium"),
)
# Right-edge cell (E) of merged D:E — only right/top/bottom needed
_HDR_BORDER_E = Border(
    right=_side("medium"), top=_side("medium"), bottom=_side("medium"),
)
_DAY_BORDER_A = Border(
    left=_side("medium"), right=_side("thin"),
    top=_side("thin"), bottom=_side("thin"),
)
_DAY_BORDER_B = Border(
    left=_side("thin"), right=_side("thin"),
    top=_side("thin"), bottom=_side("thin"),
)
_DAY_BORDER_C = Border(
    left=_side("thin"), right=_side("thin"),
    top=_side("thin"), bottom=_side("thin"),
)
_DAY_BORDER_D = Border(
    left=_side("thin"), right=_side("medium"),
    top=_side("thin"), bottom=_side("thin"),
)
# Right-edge cell (E) of merged D:E for day rows
_DAY_BORDER_E = Border(
    right=_side("medium"), top=_side("thin"), bottom=_side("thin"),
)
_TOTAL_BORDER_A = Border(
    left=_side("medium"), right=_side("thin"),
    top=_side("medium"), bottom=_side("medium"),
)
_TOTAL_BORDER_B = Border(
    left=_side("thin"), right=_side("thin"),
    top=_side("medium"), bottom=_side("medium"),
)
_TOTAL_BORDER_C = Border(
    left=_side("thin"), right=_side("thin"),
    top=_side("medium"), bottom=_side("medium"),
)
_TOTAL_BORDER_D = Border(
    left=_side("thin"), right=_side("medium"),
    top=_side("medium"), bottom=_side("medium"),
)
_TOTAL_BORDER_E = Border(
    right=_side("medium"), top=_side("medium"), bottom=_side("medium"),
)

# Info block (rows 3-8) borders — matched to sample vykaz.xlsx
# Row 3: title, all-medium box
_INFO_TITLE_A = Border(
    left=_side("medium"), right=_side("medium"),
    top=_side("medium"), bottom=_side("medium"),
)
_INFO_TITLE_MID = Border(top=_side("medium"), bottom=_side("medium"))
_INFO_TITLE_E = Border(right=_side("medium"), top=_side("medium"), bottom=_side("medium"))
# Rows 4-6: label | spacer-B | spacer-C | value | right-edge
_INFO_LABEL = Border(
    left=_side("medium"), right=_side("thin"),
    top=_side("thin"), bottom=_side("thin"),
)
_INFO_MID_B = Border(top=_side("thin"), bottom=_side("thin"))
_INFO_MID_C = Border(right=_side("thin"), top=_side("thin"), bottom=_side("thin"))
_INFO_VALUE_D = Border(
    left=_side("thin"), right=_side("medium"),
    top=_side("thin"), bottom=_side("thin"),
)
_INFO_VALUE_E = Border(right=_side("medium"), top=_side("thin"), bottom=_side("thin"))
# Row 7: empty spacer
_INFO_SPACER_A = Border(left=_side("medium"), top=_side("thin"))
_INFO_SPACER_B = Border(top=_side("thin"))
_INFO_SPACER_C = Border(right=_side("thin"), top=_side("thin"))
_INFO_SPACER_D = Border(
    left=_side("thin"), right=_side("medium"),
    top=_side("thin"), bottom=_side("thin"),
)
_INFO_SPACER_E = Border(right=_side("medium"), top=_side("thin"), bottom=_side("thin"))
# Row 8: month / year — bottom=medium acts as separator before column headers
_INFO_MONTH_A = Border(
    left=_side("medium"), right=_side("thin"),
    top=_side("thin"), bottom=_side("medium"),
)
_INFO_MONTH_B = Border(right=_side("thin"), top=_side("thin"), bottom=_side("medium"))
_INFO_MONTH_C = Border(
    left=_side("thin"), right=_side("thin"),
    top=_side("thin"), bottom=_side("medium"),
)
_INFO_MONTH_D = Border(
    left=_side("thin"), right=_side("thin"),
    top=_side("thin"), bottom=_side("medium"),
)
_INFO_MONTH_E = Border(
    left=_side("thin"), right=_side("medium"),
    top=_side("thin"), bottom=_side("medium"),
)

# ── Layout constants ──────────────────────────────────────────────────────────

_ROW_HEIGHT = 15.75
_COL_WIDTHS = {
    "A": 6.86,   # date number
    "B": 6.43,   # day abbreviation
    "C": 9.43,   # hours
    "D": 20.29,  # description (merged D:E)
    "E": 22.43,
}

_FIRST_DATA_ROW = 10  # row index of day-1 (1-based)
_HEADER_ROW = 9  # column header row


# ── Helpers ───────────────────────────────────────────────────────────────────

def _apply_row_height(ws: Worksheet, row: int, height: float = _ROW_HEIGHT) -> None:
    ws.row_dimensions[row].height = height


def _write_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: object = None,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
    number_format: str | None = None,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def _fetch_events_for_month(
    user_id: str, year: int, month: int
) -> dict[int, tuple[Decimal, list[str]]]:
    """Return {day: (total_hours, [event_names])} for the user's paid completed events."""
    from app.extensions import db
    from app.models import Assignment, Event, EventSpot, EventStatus

    period_start = datetime(year, month, 1, tzinfo=timezone.utc)
    last_day = calendar.monthrange(year, month)[1]
    period_end = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)

    rows = db.session.execute(
        sa.select(Event)
        .join(EventSpot, EventSpot.event_id == Event.id)
        .join(Assignment, Assignment.spot_id == EventSpot.id)
        .where(
            Assignment.user_id == user_id,
            Event.status == EventStatus.COMPLETED,
            Event.paid.is_(True),
            Event.start_datetime >= period_start,
            Event.start_datetime <= period_end,
        )
    ).scalars().all()

    result: dict[int, tuple[Decimal, list[str]]] = {}
    for ev in rows:
        hours = ev.billable_hours
        day = ev.start_datetime.day
        if day in result:
            prev_hours, prev_names = result[day]
            result[day] = (prev_hours + hours, prev_names + [ev.name])
        else:
            result[day] = (hours, [ev.name])
    return result


# ── Core generator ────────────────────────────────────────────────────────────

def generate_work_report(user: UserAccount, year: int, month: int) -> Path:
    """
    Build the výkaz práce xlsx for *user* for the given *year*/*month*.

    The file is written to  instance/work_report/<user_id>/<year>-<MM>.xlsx
    and is overwritten if it already exists.  Returns the absolute Path.
    """

    month_name = CZ_MONTH_NAMES[month]
    days_in_month = calendar.monthrange(year, month)[1]
    cz_holidays: set[date] = set(holidays.CZ(years=year).keys())

    events_by_day = _fetch_events_for_month(str(user.id), year, month)

    # ── Create workbook ───────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = month_name

    # Column widths
    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Page setup: landscape, A4, fit to 1 page
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.787
    ws.page_margins.bottom = 0.787
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0

    # ── Header block (rows 1-8) ───────────────────────────────────────────────
    for r in range(1, 9):
        _apply_row_height(ws, r)

    # Row 3: title (merged A3:E3) — medium border all round, cyan fill
    ws.merge_cells("A3:E3")
    _write_cell(ws, 3, 1, "Výkaz práce / odpracovaných hodin",
                font=_BOLD_FONT, fill=_CYAN_FILL,
                alignment=Alignment(horizontal="center"),
                border=_INFO_TITLE_A)
    for col in (2, 3, 4):
        _write_cell(ws, 3, col, None, border=_INFO_TITLE_MID)
    _write_cell(ws, 3, 5, None, border=_INFO_TITLE_E)

    # Rows 4-6: label (A, cyan) | B | C | value (D:E)
    for row_num, label, value in (
        (4, "Jméno pracovníka:", user.name),
        (5, "Pracovní úvazek:", "DPP"),
        (6, "pozice:", "zdravotní dozory"),
    ):
        ws.merge_cells(f"A{row_num}:C{row_num}")
        ws.merge_cells(f"D{row_num}:E{row_num}")
        _write_cell(ws, row_num, 1, label, font=_STD_FONT,
                    fill=_CYAN_FILL, border=_INFO_LABEL)
        _write_cell(ws, row_num, 2, None, border=_INFO_MID_B)
        _write_cell(ws, row_num, 3, None, border=_INFO_MID_C)
        _write_cell(ws, row_num, 4, value, font=_STD_FONT, border=_INFO_VALUE_D)
        _write_cell(ws, row_num, 5, None, border=_INFO_VALUE_E)

    # Row 7: empty spacer with side borders
    _write_cell(ws, 7, 1, None, border=_INFO_SPACER_A)
    _write_cell(ws, 7, 2, None, border=_INFO_SPACER_B)
    _write_cell(ws, 7, 3, None, border=_INFO_SPACER_C)
    _write_cell(ws, 7, 4, None, border=_INFO_SPACER_D)
    _write_cell(ws, 7, 5, None, border=_INFO_SPACER_E)

    # Row 8: month + year — bottom=medium separator before column headers
    _write_cell(ws, 8, 1, "Měsíc:", font=_STD_FONT,
                fill=_CYAN_FILL, border=_INFO_MONTH_A)
    _write_cell(ws, 8, 2, None, fill=_CYAN_FILL, border=_INFO_MONTH_B)
    _write_cell(ws, 8, 3, month_name, font=_STD_FONT,
                alignment=Alignment(horizontal="left"), border=_INFO_MONTH_C)
    _write_cell(ws, 8, 4, "Rok:", font=_STD_FONT,
                fill=_CYAN_FILL, border=_INFO_MONTH_D)
    _write_cell(ws, 8, 5, year, font=_STD_FONT,
                alignment=Alignment(horizontal="left"), border=_INFO_MONTH_E)

    # ── Column headers (row 9) ────────────────────────────────────────────────
    _apply_row_height(ws, _HEADER_ROW)
    ws.merge_cells("D9:E9")
    _write_cell(ws, 9, 1, "Datum",
                font=_BOLD_FONT, fill=_BLUE_FILL,
                alignment=Alignment(horizontal="center"),
                border=_HDR_BORDER_A)
    _write_cell(ws, 9, 2, "Den",
                font=_BOLD_FONT, fill=_BLUE_FILL,
                alignment=Alignment(horizontal="center"),
                border=_HDR_BORDER_B)
    _write_cell(ws, 9, 3, "Počet hodin",
                font=_BOLD_FONT, fill=_BLUE_FILL,
                alignment=Alignment(horizontal="center", wrap_text=True),
                border=_HDR_BORDER_C)
    _write_cell(ws, 9, 4, "Popis činnosti",
                font=_BOLD_FONT, fill=_BLUE_FILL,
                alignment=Alignment(horizontal="center"),
                border=_HDR_BORDER_D)
    # Right edge of merged D9:E9
    _write_cell(ws, 9, 5, None, border=_HDR_BORDER_E)

    # ── Day rows ──────────────────────────────────────────────────────────────
    for day in range(1, days_in_month + 1):
        row = _FIRST_DATA_ROW + day - 1
        _apply_row_height(ws, row)

        d = date(year, month, day)
        weekday = d.weekday()  # 0=Mon … 6=Sun
        is_weekend = weekday >= 5  # Sat or Sun
        is_holiday = d in cz_holidays

        fill = _YELLOW_FILL if is_holiday else _WHITE_FILL
        day_font = _RED_FONT if is_weekend else _STD_FONT

        hours_val, names = events_by_day.get(day, (None, []))
        hours_display = float(hours_val) if hours_val else None
        description = ", ".join(names) if names else None

        ws.merge_cells(f"D{row}:E{row}")

        _write_cell(ws, row, 1, day,
                    font=_STD_FONT, fill=fill,
                    alignment=Alignment(horizontal="center"),
                    border=_DAY_BORDER_A)
        _write_cell(ws, row, 2, CZ_WEEKDAY_ABBR[weekday],
                    font=day_font,
                    alignment=Alignment(horizontal="center"),
                    border=_DAY_BORDER_B)
        _write_cell(ws, row, 3, hours_display,
                    font=_STD_FONT,
                    alignment=Alignment(horizontal="center"),
                    border=_DAY_BORDER_C)
        _write_cell(ws, row, 4, description,
                    font=_STD_FONT,
                    alignment=Alignment(horizontal="left"),
                    border=_DAY_BORDER_D)
        # Right edge of merged D:E
        _write_cell(ws, row, 5, None, border=_DAY_BORDER_E)

    # ── Totals row ────────────────────────────────────────────────────────────
    total_row = _FIRST_DATA_ROW + days_in_month
    _apply_row_height(ws, total_row)
    total_hours = float(sum(h for h, _ in events_by_day.values())) if events_by_day else 0.0
    ws.merge_cells(f"D{total_row}:E{total_row}")
    _write_cell(ws, total_row, 1, "Celkem hodin",
                font=_BOLD_FONT, border=_TOTAL_BORDER_A)
    _write_cell(ws, total_row, 2, None, font=_BOLD_FONT,
                border=_TOTAL_BORDER_B)
    _write_cell(ws, total_row, 3, total_hours,
                font=_BOLD_FONT,
                alignment=Alignment(horizontal="center"),
                border=_TOTAL_BORDER_C)
    _write_cell(ws, total_row, 4, None, border=_TOTAL_BORDER_D)
    # Right edge of merged D:E
    _write_cell(ws, total_row, 5, None, border=_TOTAL_BORDER_E)

    # ── Signature rows ────────────────────────────────────────────────────────
    sig_worker = total_row + 4
    sig_boss = total_row + 7
    for r in (sig_worker, sig_boss):
        _apply_row_height(ws, r)

    _write_cell(ws, sig_worker, 1, "Datum a podpis pracovníka:", font=_BOLD_FONT)
    last_day_date = date(year, month, days_in_month)
    _write_cell(ws, sig_worker, 4, last_day_date, font=_STD_FONT,
                number_format="DD.MM.YYYY")
    _write_cell(ws, sig_boss, 1, "Datum a podpis nadřízeného pracovníka:", font=_BOLD_FONT)

    # ── Save ──────────────────────────────────────────────────────────────────
    instance_path = Path(current_app.instance_path)
    out_dir = instance_path / "work_report" / str(user.id)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{year}-{month:02d}.xlsx"

    wb.save(str(out_path))
    return out_path
